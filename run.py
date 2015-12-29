#coding=utf-8
import sys
from openpyxl import workbook
from openpyxl import load_workbook
import sys
import getpage
reload(sys)
sys.setdefaultencoding('utf-8')
file_name = r'C:/Users/samsung/Desktop/s.xlsx'
#u_file_name = unicode(file_name,'utf-8')
#print u_file_name
#with load_workbook(filename=file_name) as f1:
s = u'微博活动一参与者汇总'
books = load_workbook(filename=file_name)
book = books.get_sheet_by_name(s)
request = getpage.http_page('http://weibo.com/u/')
for r_index in xrange(len(book.rows)):
	if r_index<10:
		#if book.cell(row=r_index+1,column=5).value == 'NULL':
		request.get(book.cell(row=r_index+2,column=1).value)
		#book.cell(row=r_index+1,column=5).value = 123
books.save(filename=file_name)
 
#print type(book.rows)
#print s
#print len(book.rows)